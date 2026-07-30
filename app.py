from flask import Flask, request, jsonify, send_file, render_template, redirect, url_for, make_response
from flask_cors import CORS
from flask_apscheduler import APScheduler
import sqlite3
from datetime import datetime, timedelta
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import requests
import json
import logging
import os

app = Flask(__name__)
app.config['SECRET_KEY'] = 'guankong_secret_key_2024'
app.config['DEBUG'] = True
app.config['TEMPLATES_AUTO_RELOAD'] = True
CORS(app)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

DATABASE = 'guankong.db'
SCHEDULER_API_ENABLED = True

scheduler = APScheduler()
scheduler.init_app(app)
scheduler.start()

class MockUser:
    id = 1
    username = 'admin'
    role = '管理员'
    department = '管理部'

current_user = MockUser()

def parse_date(date_str):
    if date_str is None:
        return datetime.now().strftime('%Y-%m-%d')
    
    if isinstance(date_str, datetime):
        return date_str.strftime('%Y-%m-%d')
    
    date_str = str(date_str).strip()
    if not date_str:
        return datetime.now().strftime('%Y-%m-%d')
    
    formats = [
        '%Y-%m-%d',
        '%Y.%m.%d',
        '%Y/%m/%d',
        '%Y-%m-%d %H:%M:%S',
        '%Y.%m.%d %H:%M:%S',
        '%Y/%m/%d %H:%M:%S',
        '%m/%d/%Y',
        '%d/%m/%Y',
        '%Y年%m月%d日',
        '%Y/%m/%d',
        '%d-%m-%Y',
        '%d.%m.%Y',
        '%m-%d-%Y'
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    
    return datetime.now().strftime('%Y-%m-%d')

def init_db():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT '操作员',
            department TEXT NOT NULL DEFAULT '仓储',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    cursor.execute('SELECT COUNT(*) FROM users')
    if cursor.fetchone()[0] == 0:
        cursor.execute('INSERT INTO users (username, password, role, department) VALUES (?, ?, ?, ?)',
                      ('admin', 'admin123', '管理员', '管理部'))
        cursor.execute('INSERT INTO users (username, password, role, department) VALUES (?, ?, ?, ?)',
                      ('user001', '123456', '操作员', '仓储部'))
        cursor.execute('INSERT INTO users (username, password, role, department) VALUES (?, ?, ?, ?)',
                      ('user002', '654321', '操作员', '质量部'))
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS guankong_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            production_date DATE,
            product_name TEXT,
            control_reason TEXT,
            quantity INTEGER,
            handle_opinion TEXT,
            handle_time DATE,
            handle_dept TEXT,
            status TEXT DEFAULT '待处理',
            deadline DATE,
            remark TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            handled_quantity INTEGER DEFAULT 0,
            creator_id INTEGER,
            handler_name TEXT
        )
    ''')
    
    cursor.execute('PRAGMA table_info(guankong_records)')
    columns = [col[1] for col in cursor.fetchall()]
    if 'handler_name' not in columns:
        cursor.execute('ALTER TABLE guankong_records ADD COLUMN handler_name TEXT')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS system_config (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            config_key TEXT UNIQUE NOT NULL,
            config_value TEXT NOT NULL,
            description TEXT
        )
    ''')
    
    cursor.execute('SELECT COUNT(*) FROM system_config')
    if cursor.fetchone()[0] == 0:
        cursor.execute('INSERT INTO system_config (config_key, config_value, description) VALUES (?, ?, ?)',
                      ('default_deadline_days', '7', '默认处理期限天数'))
        cursor.execute('INSERT INTO system_config (config_key, config_value, description) VALUES (?, ?, ?)',
                      ('overdue_remind_days', '5', '超期提醒天数阈值'))
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS feishu_config (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            webhook_url TEXT NOT NULL DEFAULT '',
            external_url TEXT NOT NULL DEFAULT 'http://localhost:5000',
            enabled INTEGER NOT NULL DEFAULT 1,
            schedule_time TEXT NOT NULL DEFAULT '08:00',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    cursor.execute('SELECT COUNT(*) FROM feishu_config')
    if cursor.fetchone()[0] == 0:
        cursor.execute('''
            INSERT INTO feishu_config (webhook_url, external_url, enabled, schedule_time) 
            VALUES (?, ?, ?, ?)
        ''', ('', 'http://localhost:5000', 1, '08:00'))
    
    conn.commit()
    conn.close()

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def get_config(key):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT config_value FROM system_config WHERE config_key = ?', (key,))
    row = cursor.fetchone()
    conn.close()
    return row['config_value'] if row else None

def set_config(key, value):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE system_config SET config_value = ? WHERE config_key = ?', (value, key))
    conn.commit()
    conn.close()

def get_feishu_config():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM feishu_config LIMIT 1')
    row = cursor.fetchone()
    conn.close()
    if row:
        return {
            'id': row['id'],
            'webhook_url': row['webhook_url'],
            'external_url': row['external_url'],
            'enabled': bool(row['enabled']),
            'schedule_time': row['schedule_time']
        }
    return None

def save_feishu_config(config):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE feishu_config 
        SET webhook_url = ?, external_url = ?, enabled = ?, schedule_time = ?, updated_at = ?
        WHERE id = ?
    ''', (config['webhook_url'], config['external_url'], int(config['enabled']), 
          config['schedule_time'], datetime.now().strftime('%Y-%m-%d %H:%M:%S'), config['id']))
    conn.commit()
    conn.close()

def send_feishu_card(report_data):
    config = get_feishu_config()
    if not config or not config['enabled'] or not config['webhook_url']:
        logger.warning("飞书通知未启用或未配置Webhook")
        return False, "飞书通知未启用或未配置Webhook"
    
    try:
        card = build_feishu_card(report_data, config['external_url'])
        response = requests.post(config['webhook_url'], json=card)
        if response.status_code == 200:
            logger.info(f"飞书卡片通知发送成功")
            return True, "发送成功"
        else:
            logger.error(f"飞书卡片通知发送失败: {response.status_code} - {response.text}")
            return False, f"发送失败: {response.text}"
    except Exception as e:
        logger.error(f"飞书卡片通知发送失败: {str(e)}")
        return False, str(e)

def build_feishu_card(report_data, external_url):
    today = datetime.now().strftime('%Y-%m-%d')
    title = f"📦 管制品处理进度通报 ({today})"
    
    overdue_count = report_data.get('overdue_count', 0)
    if overdue_count <= 3:
        template_color = "blue"
    elif overdue_count <= 7:
        template_color = "orange"
    else:
        template_color = "red"
    
    pending_qty = report_data['pending_qty']
    processing_qty = report_data['processing_qty']
    completed_qty = report_data['completed_qty']
    total_qty = pending_qty + processing_qty + completed_qty
    completion_rate = round((completed_qty / total_qty) * 100) if total_qty > 0 else 0
    
    header = {
        "title": {"tag": "plain_text", "content": title},
        "template": template_color
    }
    
    elements = []
    
    stats_text = f"**📊 统计概览**\n\n"
    stats_text += f"| 状态 | 数量 |\n| --- | --- |\n"
    stats_text += f"| 🔴待处理 | {pending_qty:,}箱 |\n"
    stats_text += f"| 🟠处理中 | {processing_qty:,}箱 |\n"
    stats_text += f"| 🟢已完成 | {completed_qty:,}箱 |\n"
    stats_text += f"| ⚠️超期 | {overdue_count}项 |\n\n"
    stats_text += f"**整体完成率**: {completion_rate}%\n"
    
    stats_section = {
        "tag": "div",
        "text": {"tag": "lark_md", "content": stats_text}
    }
    elements.append(stats_section)
    
    if report_data.get('has_overdue', True) and report_data['overdue_records']:
        overdue_text = "\n**🔴 超期未处理 TOP 5**\n\n"
        for record in report_data.get('overdue_records', [])[:5]:
            urgency_icon = "🔥" if record['days_overdue'] > 10 else "⚠️" if record['days_overdue'] > 5 else "⏰"
            overdue_text += f"{urgency_icon} {record['product_name']}\n"
            overdue_text += f"  - 原因: {record['control_reason']}\n"
            overdue_text += f"  - 数量: {record['quantity']}箱\n"
            overdue_text += f"  - 超期: {record['days_overdue']}天\n\n"
        
        overdue_section = {
            "tag": "div",
            "text": {"tag": "lark_md", "content": overdue_text}
        }
        elements.append(overdue_section)
    else:
        no_overdue = {
            "tag": "div",
            "text": {"tag": "lark_md", "content": "\n🎉 **今日无超期管制品，请继续保持！**"}
        }
        elements.append(no_overdue)
    
    button_action = {
        "tag": "action",
        "actions": [
            {
                "tag": "button",
                "text": {"tag": "plain_text", "content": "📊 查看全部进度"},
                "type": "primary",
                "url": f"{external_url}/dashboard"
            }
        ],
        "layout": "center"
    }
    elements.append(button_action)
    
    card = {
        "msg_type": "interactive",
        "card": {
            "header": header,
            "elements": elements,
            "config": {"wide_screen_mode": True},
            "footer": {
                "text": {"tag": "plain_text", "content": "管制品追踪系统 · 每日自动推送"}
            }
        }
    }
    
    return card

@app.route('/api/feishu/config', methods=['GET'])
def get_feishu_config_api():
    config = get_feishu_config()
    if config:
        return jsonify(config)
    return jsonify({'error': '配置不存在'}), 404

@app.route('/api/feishu/config', methods=['POST'])
def save_feishu_config_api():
    data = request.json
    config = get_feishu_config()
    if config:
        data['id'] = config['id']
        save_feishu_config(data)
        return jsonify({'message': '配置保存成功'})
    return jsonify({'error': '配置不存在'}), 404

@app.route('/api/feishu/test', methods=['POST'])
def test_feishu_card():
    today = datetime.now().strftime('%Y-%m-%d')
    test_data = {
        'pending_qty': 100,
        'processing_qty': 50,
        'completed_qty': 200,
        'overdue_count': 3,
        'has_overdue': True,
        'overdue_records': [
            {'product_name': '测试产品A', 'control_reason': '质量问题', 'quantity': 50, 'days_overdue': 2},
            {'product_name': '测试产品B', 'control_reason': '超期', 'quantity': 30, 'days_overdue': 5}
        ]
    }
    success, msg = send_feishu_card(test_data)
    return jsonify({'success': success, 'message': msg})

@app.route('/api/feishu/send-report', methods=['POST'])
def send_feishu_report():
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT COALESCE(SUM(quantity), 0) FROM guankong_records WHERE status = "待处理"')
    pending_qty = cursor.fetchone()[0]
    
    cursor.execute('SELECT COALESCE(SUM(quantity), 0) FROM guankong_records WHERE status = "处理中"')
    processing_qty = cursor.fetchone()[0]
    
    cursor.execute('SELECT COALESCE(SUM(quantity), 0) FROM guankong_records WHERE status = "已完成"')
    completed_qty = cursor.fetchone()[0]
    
    today = datetime.now().strftime('%Y-%m-%d')
    cursor.execute('SELECT COUNT(*) FROM guankong_records WHERE status != "已完成" AND deadline < ?', (today,))
    overdue_count = cursor.fetchone()[0]
    
    cursor.execute('''
        SELECT product_name, control_reason, quantity, deadline 
        FROM guankong_records 
        WHERE status != "已完成" AND deadline < ?
        ORDER BY deadline ASC
    ''', (today,))
    overdue_records = []
    for row in cursor.fetchall():
        days_overdue = (datetime.now() - datetime.strptime(row['deadline'], '%Y-%m-%d')).days
        overdue_records.append({
            'product_name': row['product_name'],
            'control_reason': row['control_reason'],
            'quantity': row['quantity'],
            'days_overdue': days_overdue
        })
    
    conn.close()
    
    report_data = {
        'pending_qty': pending_qty,
        'processing_qty': processing_qty,
        'completed_qty': completed_qty,
        'overdue_count': overdue_count,
        'has_overdue': overdue_count > 0,
        'overdue_records': overdue_records
    }
    
    success, msg = send_feishu_card(report_data)
    return jsonify({'success': success, 'message': msg})

@app.route('/api/feishu/notify-overdue', methods=['POST'])
def notify_overdue():
    conn = get_db()
    cursor = conn.cursor()
    
    today = datetime.now().strftime('%Y-%m-%d')
    cursor.execute('''
        SELECT product_name, handle_dept, quantity, deadline 
        FROM guankong_records 
        WHERE status = "待处理" AND (julianday('now') - julianday(deadline)) > 0
        ORDER BY (julianday('now') - julianday(deadline)) DESC
        LIMIT 10
    ''')
    
    overdue_records = []
    for row in cursor.fetchall():
        days_overdue = (datetime.now() - datetime.strptime(row['deadline'], '%Y-%m-%d')).days
        overdue_records.append({
            'product_name': row['product_name'],
            'handle_dept': row['handle_dept'],
            'quantity': row['quantity'],
            'days_overdue': days_overdue
        })
    
    conn.close()
    
    notify_data = {
        'pending_qty': 0,
        'processing_qty': 0,
        'completed_qty': 0,
        'overdue_count': len(overdue_records),
        'has_overdue': len(overdue_records) > 0,
        'overdue_records': overdue_records
    }
    
    success, msg = send_feishu_card(notify_data)
    return jsonify({'success': success, 'message': msg})

@scheduler.task('cron', id='daily_feishu_report', hour=8, minute=0)
def daily_feishu_report():
    config = get_feishu_config()
    if not config or not config['enabled']:
        return
    
    weekday = datetime.now().weekday()
    if weekday >= 5:
        return
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT COALESCE(SUM(quantity), 0) FROM guankong_records WHERE status = "待处理"')
    pending_qty = cursor.fetchone()[0]
    
    cursor.execute('SELECT COALESCE(SUM(quantity), 0) FROM guankong_records WHERE status = "处理中"')
    processing_qty = cursor.fetchone()[0]
    
    cursor.execute('SELECT COALESCE(SUM(quantity), 0) FROM guankong_records WHERE status = "已完成"')
    completed_qty = cursor.fetchone()[0]
    
    today = datetime.now().strftime('%Y-%m-%d')
    cursor.execute('SELECT COUNT(*) FROM guankong_records WHERE status != "已完成" AND deadline < ?', (today,))
    overdue_count = cursor.fetchone()[0]
    
    cursor.execute('''
        SELECT product_name, control_reason, quantity, deadline 
        FROM guankong_records 
        WHERE status != "已完成" AND deadline < ?
        ORDER BY deadline ASC
    ''', (today,))
    overdue_records = []
    for row in cursor.fetchall():
        days_overdue = (datetime.now() - datetime.strptime(row['deadline'], '%Y-%m-%d')).days
        overdue_records.append({
            'product_name': row['product_name'],
            'control_reason': row['control_reason'],
            'quantity': row['quantity'],
            'days_overdue': days_overdue
        })
    
    conn.close()
    
    report_data = {
        'pending_qty': pending_qty,
        'processing_qty': processing_qty,
        'completed_qty': completed_qty,
        'overdue_count': overdue_count,
        'has_overdue': overdue_count > 0,
        'overdue_records': overdue_records
    }
    
    send_feishu_card(report_data)

@app.route('/api/send_overdue_reminder', methods=['POST'])
def send_overdue_reminder():
    conn = get_db()
    cursor = conn.cursor()
    
    overdue_days = int(get_config('overdue_remind_days') or 5)
    threshold_date = (datetime.now() - timedelta(days=overdue_days)).strftime('%Y-%m-%d')
    
    cursor.execute('''
        SELECT product_name, control_reason, quantity, handle_dept, deadline, handler_name 
        FROM guankong_records 
        WHERE status != "已完成" AND deadline < ?
        ORDER BY deadline DESC
    ''', (threshold_date,))
    
    records = cursor.fetchall()
    conn.close()
    
    if records:
        title = f"⚠️ 超期提醒：{len(records)}项管制品超期超过{overdue_days}天"
        send_feishu_card_notification(title, records)
        return jsonify({'message': f'已向飞书推送 {len(records)} 条超期提醒', 'count': len(records)})
    else:
        return jsonify({'message': '没有超期超过5天的管制品', 'count': 0})

@app.route('/')
def index():
    return redirect(url_for('dashboard'))

@app.route('/dashboard')
def dashboard():
    response = make_response(render_template('dashboard.html'))
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = 'Thu, 01 Jan 1970 00:00:00 GMT'
    return response

@app.route('/mobile')
def mobile():
    return render_template('mobile.html')

@app.route('/records')
def records():
    return render_template('records.html')

@app.route('/add')
def add():
    return render_template('add.html')

@app.route('/import')
def import_page():
    return render_template('import.html')

@app.route('/report')
def report():
    return render_template('report.html')

@app.route('/admin')
def admin():
    return render_template('admin.html')

@app.route('/api/statistics')
def get_statistics():
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT COALESCE(SUM(quantity), 0) FROM guankong_records')
    total_qty_all = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM guankong_records')
    pending_count = cursor.fetchone()[0]
    
    today = datetime.now().strftime('%Y-%m-%d')
    today_added = 0
    
    cursor.execute('SELECT COALESCE(SUM(quantity), 0) FROM guankong_records WHERE status != "已完成" AND deadline < ?', (today,))
    overdue_qty = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM guankong_records WHERE status != "已完成" AND deadline < ?', (today,))
    overdue_count = cursor.fetchone()[0]
    
    cursor.execute('SELECT COALESCE(SUM(quantity), 0) FROM guankong_records WHERE status != "已完成" AND deadline < ?', (today,))
    today_due_qty = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM guankong_records WHERE status != "已完成" AND deadline < ?', (today,))
    today_due_count = cursor.fetchone()[0]
    
    cursor.execute('SELECT COALESCE(SUM(quantity), 0) FROM guankong_records')
    total_qty = cursor.fetchone()[0]
    
    cursor.execute('SELECT COALESCE(SUM(quantity), 0) FROM guankong_records WHERE status = "已完成"')
    completed_qty = cursor.fetchone()[0]
    
    cursor.execute('SELECT COALESCE(SUM(quantity), 0) FROM guankong_records WHERE status != "已完成"')
    uncompleted_qty = cursor.fetchone()[0]
    
    completion_rate = round((completed_qty / total_qty) * 100) if total_qty > 0 else 0
    
    cursor.execute('''
        SELECT COALESCE(handler_name, handle_dept) as dept,
               COUNT(*) as count,
               COALESCE(SUM(quantity), 0) as quantity,
               COALESCE(SUM(CASE WHEN (julianday('now') - julianday(deadline)) > 15 THEN quantity ELSE 0 END), 0) as critical_overdue,
               COALESCE(SUM(CASE WHEN (julianday('now') - julianday(deadline)) BETWEEN 7 AND 15 THEN quantity ELSE 0 END), 0) as normal_overdue,
               COALESCE(SUM(CASE WHEN (julianday('now') - julianday(deadline)) BETWEEN 5 AND 7 THEN quantity ELSE 0 END), 0) as upcoming_due,
               COALESCE(SUM(CASE WHEN (julianday('now') - julianday(deadline)) < 5 THEN quantity ELSE 0 END), 0) as normal
        FROM guankong_records
        WHERE status = "待处理"
          AND deadline IS NOT NULL
          AND COALESCE(handler_name, handle_dept) IS NOT NULL
          AND COALESCE(handler_name, handle_dept) != ""
          AND COALESCE(handler_name, handle_dept) != "未分配"
        GROUP BY COALESCE(handler_name, handle_dept)
        ORDER BY SUM(quantity) DESC
    ''')
    dept_stats = []
    for row in cursor.fetchall():
        dept_stats.append({
            'dept': row['dept'],
            'count': row['count'],
            'qty': row['quantity'],
            'value': row['quantity'],
            'quantity': row['quantity'],
            'critical_overdue': row['critical_overdue'],
            'normal_overdue': row['normal_overdue'],
            'upcoming_due': row['upcoming_due'],
            'normal': row['normal']
        })
    
    cursor.execute('''
        SELECT COALESCE(handler_name, handle_dept) as dept, 
               COALESCE(SUM(CASE WHEN status != "已完成" AND deadline < ? THEN quantity ELSE 0 END), 0) as overdue_qty,
               COALESCE(AVG(JULIANDAY(?) - JULIANDAY(production_date)), 0) as avg_wait_days
        FROM guankong_records 
        WHERE status = "待处理"
          AND COALESCE(handler_name, handle_dept) IS NOT NULL
          AND COALESCE(handler_name, handle_dept) != ""
          AND COALESCE(handler_name, handle_dept) != "未分配"
        GROUP BY COALESCE(handler_name, handle_dept)
    ''', (today, today))
    for row in cursor.fetchall():
        for stat in dept_stats:
            if stat['dept'] == row['dept']:
                stat['overdue_qty'] = row['overdue_qty']
                stat['avg_wait_days'] = round(row['avg_wait_days'])
                break
    
    cursor.execute('SELECT handle_dept, COALESCE(SUM(quantity), 0) FROM guankong_records WHERE status != "已完成" AND deadline < ? GROUP BY handle_dept', (today,))
    overdue_dept_qty = {}
    for row in cursor.fetchall():
        overdue_dept_qty[row['handle_dept']] = row[1]
    
    cursor.execute('SELECT handle_dept, COALESCE(SUM(quantity), 0) FROM guankong_records WHERE status = "已完成" GROUP BY handle_dept')
    completed_dept_qty = {}
    total_dept_qty = {}
    for row in cursor.fetchall():
        completed_dept_qty[row['handle_dept']] = row[1]
    
    cursor.execute('SELECT handle_dept, COALESCE(SUM(quantity), 0) FROM guankong_records GROUP BY handle_dept')
    for row in cursor.fetchall():
        total_dept_qty[row['handle_dept']] = row[1]
    
    cursor.execute('SELECT DISTINCT COALESCE(handler_name, handle_dept) as dept FROM guankong_records WHERE COALESCE(handler_name, handle_dept) IS NOT NULL AND COALESCE(handler_name, handle_dept) != "" AND COALESCE(handler_name, handle_dept) != "未分配"')
    all_depts = [row['dept'] for row in cursor.fetchall()]
    
    dept_percentages = {
        'pending': {},
        'overdue': {},
        'completed': {}
    }
    for dept in all_depts:
        dept_percentages['pending'][dept] = 0
        dept_percentages['overdue'][dept] = 0
        dept_percentages['completed'][dept] = 0
    
    if uncompleted_qty > 0:
        cursor.execute('SELECT COALESCE(handler_name, handle_dept) as dept, COALESCE(SUM(quantity), 0) FROM guankong_records WHERE status = "待处理" GROUP BY COALESCE(handler_name, handle_dept)')
        pending_dept = {}
        for row in cursor.fetchall():
            pending_dept[row['dept']] = row[1]
        for dept in all_depts:
            dept_percentages['pending'][dept] = round((pending_dept.get(dept, 0) / uncompleted_qty) * 100)
    
    if overdue_qty > 0:
        for dept in all_depts:
            dept_percentages['overdue'][dept] = round((overdue_dept_qty.get(dept, 0) / overdue_qty) * 100)
    
    for dept in all_depts:
        total = total_dept_qty.get(dept, 0)
        if total > 0:
            dept_percentages['completed'][dept] = round((completed_dept_qty.get(dept, 0) / total) * 100)
    
    last_7_days = []
    for i in range(6, -1, -1):
        date = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
        
        cursor.execute('SELECT COALESCE(SUM(quantity), 0) FROM guankong_records WHERE DATE(production_date) = ?', (date,))
        created_qty = cursor.fetchone()[0]
        
        cursor.execute('SELECT COALESCE(SUM(quantity), 0) FROM guankong_records WHERE DATE(handle_time) = ?', (date,))
        completed_qty = cursor.fetchone()[0]
        
        cursor.execute('''
            SELECT COALESCE(SUM(quantity), 0) FROM guankong_records 
            WHERE status = "待处理" AND (julianday(?) - julianday(deadline)) >= 7
        ''', (date,))
        pending_qty = cursor.fetchone()[0]
        
        rate = round((completed_qty / (created_qty + pending_qty)) * 100) if (created_qty + pending_qty) > 0 else 0
        
        cursor.execute('''
            SELECT product_name, 
                   COALESCE(SUM(CASE WHEN status = "已完成" THEN quantity ELSE 0 END), 0) as completed,
                   COALESCE(SUM(quantity), 0) as total
            FROM guankong_records 
            WHERE DATE(production_date) = ?
            GROUP BY product_name 
            ORDER BY total DESC 
            LIMIT 3
        ''', (date,))
        top_products = []
        for row in cursor.fetchall():
            p_rate = round((row[1] / row[2]) * 100) if row[2] > 0 else 0
            top_products.append({'name': row[0], 'rate': p_rate})
        
        last_7_days.append({
            'date': date, 
            'count': created_qty, 
            'completed': completed_qty, 
            'rate': rate,
            'completed_qty': completed_qty,
            'pending_qty': pending_qty,
            'top_products': top_products
        })
    
    cursor.execute('''
        SELECT 
            CASE 
                WHEN product_name LIKE '成品-PET%' THEN SUBSTR(product_name, 5, INSTR(SUBSTR(product_name, 5), '入') + 4)
                ELSE product_name 
            END as spec_flavor,
            COALESCE(SUM(CASE WHEN status = "待处理" THEN quantity ELSE 0 END), 0) as pending_qty,
            COALESCE(SUM(handled_quantity), 0) as handled_qty,
            COALESCE(SUM(quantity), 0) as total_qty,
            COALESCE(SUM(CASE WHEN (julianday('now') - julianday(deadline)) > 0 THEN quantity ELSE 0 END), 0) as overdue_qty,
            COALESCE(AVG(julianday('now') - julianday(production_date)), 0) as avg_wait_days,
            GROUP_CONCAT(product_name, ';') as product_names
        FROM guankong_records 
        GROUP BY spec_flavor 
        ORDER BY pending_qty DESC 
        LIMIT 10
    ''')
    product_stats = []
    for row in cursor.fetchall():
        total = row[3]
        handled = row[2]
        rate = round((handled / total) * 100) if total > 0 else 0
        product_names = row[6]
        if product_names:
            product_names = ';'.join(list(dict.fromkeys(product_names.split(';'))))
        product_stats.append({
            'spec_flavor': row[0],
            'pending_qty': row[1],
            'handled_qty': handled,
            'total_qty': total,
            'rate': rate,
            'overdue_qty': row[4],
            'avg_wait_days': round(row[5]) if row[5] else 0,
            'product_names': product_names
        })
    
    cursor.execute('SELECT control_reason, COALESCE(SUM(quantity), 0) as count FROM guankong_records GROUP BY control_reason ORDER BY SUM(quantity) DESC')
    all_reasons = []
    total_count = 0
    for row in cursor.fetchall():
        all_reasons.append({'reason': row['control_reason'], 'count': row[1]})
        total_count += row[1]
    
    reason_distribution = []
    others_count = 0
    min_percentage = 5
    
    for item in all_reasons:
        percentage = (item['count'] / total_count) * 100
        if percentage >= min_percentage:
            reason_distribution.append({'reason': item['reason'], 'count': item['count']})
        else:
            others_count += item['count']
    
    if others_count > 0:
        reason_distribution.append({'reason': '其他', 'count': others_count})
    
    conn.close()
    
    return jsonify({
        'pending_qty': total_qty_all,
        'pending_count': pending_count,
        'test_marker': 'updated_v2',
        'today_added': today_added,
        'overdue_qty': overdue_qty,
        'overdue_count': overdue_count,
        'today_due_qty': today_due_qty,
        'today_due_count': today_due_count,
        'completion_rate': completion_rate,
        'dept_stats': dept_stats,
        'dept_percentages': dept_percentages,
        'last_7_days': last_7_days,
        'product_stats': product_stats,
        'reason_distribution': reason_distribution
    })

@app.route('/api/dashboard/reason-distribution')
def get_reason_distribution():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT control_reason, SUM(quantity) as total_quantity 
        FROM guankong_records 
        WHERE status IN ("待处理", "处理中") 
        GROUP BY control_reason 
        ORDER BY total_quantity DESC
    ''')
    
    result = []
    for row in cursor.fetchall():
        result.append({
            'control_reason': row[0] if row[0] else '',
            'total_quantity': row[1]
        })
    
    conn.close()
    return app.response_class(
        response=json.dumps(result, ensure_ascii=False),
        status=200,
        mimetype='application/json; charset=utf-8'
    )

@app.route('/api/depts')
def get_depts():
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT DISTINCT handle_dept FROM guankong_records WHERE handle_dept IS NOT NULL AND handle_dept != "" ORDER BY handle_dept')
    depts = [row['handle_dept'] for row in cursor.fetchall()]
    conn.close()
    
    return jsonify(depts)

@app.route('/api/dashboard/dept-records')
def get_dept_records():
    dept = request.args.get('dept')
    conn = get_db()
    cursor = conn.cursor()
    
    if not dept:
        cursor.execute('''
            SELECT COALESCE(handler_name, handle_dept) as dept,
                   COUNT(*) as count,
                   COALESCE(SUM(quantity), 0) as quantity,
                   COALESCE(SUM(CASE WHEN (julianday('now') - julianday(deadline)) > 15 THEN quantity ELSE 0 END), 0) as critical_overdue,
                   COALESCE(SUM(CASE WHEN (julianday('now') - julianday(deadline)) BETWEEN 7 AND 15 THEN quantity ELSE 0 END), 0) as normal_overdue,
                   COALESCE(SUM(CASE WHEN (julianday('now') - julianday(deadline)) BETWEEN 5 AND 7 THEN quantity ELSE 0 END), 0) as upcoming_due,
                   COALESCE(SUM(CASE WHEN (julianday('now') - julianday(deadline)) < 5 THEN quantity ELSE 0 END), 0) as normal
            FROM guankong_records
            WHERE status = "待处理"
              AND deadline IS NOT NULL
              AND COALESCE(handler_name, handle_dept) IS NOT NULL
              AND COALESCE(handler_name, handle_dept) != ""
              AND COALESCE(handler_name, handle_dept) != "未分配"
            GROUP BY COALESCE(handler_name, handle_dept)
            ORDER BY SUM(quantity) DESC
        ''')
        records = cursor.fetchall()
        conn.close()
        return jsonify([{
            'dept': r['dept'],
            'count': r['count'],
            'quantity': r['quantity'],
            'critical_overdue': r['critical_overdue'],
            'normal_overdue': r['normal_overdue'],
            'upcoming_due': r['upcoming_due'],
            'normal': r['normal']
        } for r in records])
    
    urgency = request.args.get('urgency')
    
    query = '''SELECT product_name, quantity, deadline FROM guankong_records 
               WHERE status = "待处理" AND deadline IS NOT NULL AND (handler_name = ? OR handle_dept = ?)'''
    params = [dept, dept]
    
    if urgency:
        if urgency == '严重超期':
            query += ' AND (julianday("now") - julianday(deadline)) > 15'
        elif urgency == '一般超期':
            query += ' AND (julianday("now") - julianday(deadline)) BETWEEN 7 AND 15'
        elif urgency == '即将到期':
            query += ' AND (julianday("now") - julianday(deadline)) BETWEEN 5 AND 7'
        elif urgency == '正常':
            query += ' AND (julianday("now") - julianday(deadline)) < 5'
    
    query += ' ORDER BY deadline ASC'
    
    cursor.execute(query, params)
    records = cursor.fetchall()
    conn.close()
    
    return jsonify({'records': [{'product_name': r['product_name'], 'quantity': r['quantity'], 'deadline': r['deadline']} for r in records]})

@app.route('/api/dashboard/product-records')
def get_product_records():
    product_name = request.args.get('product_name')
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT id, production_date, product_name, control_reason, quantity, 
               handle_opinion, handle_dept, status, deadline, remark, created_at
        FROM guankong_records 
        WHERE product_name = ?
        ORDER BY production_date DESC
    ''', (product_name,))
    
    records = cursor.fetchall()
    conn.close()
    
    return jsonify({'records': [dict(r) for r in records]})

@app.route('/api/dashboard/monthly-statistics')
def get_monthly_statistics():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT handler_name, handle_dept, quantity, status, production_date 
        FROM guankong_records 
        WHERE production_date IS NOT NULL
          AND COALESCE(handler_name, handle_dept) IS NOT NULL
          AND COALESCE(handler_name, handle_dept) != ""
          AND COALESCE(handler_name, handle_dept) != "未分配"
    ''')
    
    records = cursor.fetchall()
    
    handler_data = {}
    for row in records:
        handler_name = row[0] or ''
        handle_dept = row[1] or ''
        quantity = row[2] or 0
        status = row[3] or ''
        
        handler = handler_name if handler_name else handle_dept
        
        if not handler or handler == '未分配':
            continue
        
        if handler not in handler_data:
            handler_data[handler] = {
                'total_qty': 0,
                'pending_qty': 0,
                'completed_qty': 0,
                'processing_qty': 0,
                'record_count': 0
            }
        
        handler_data[handler]['total_qty'] += quantity
        handler_data[handler]['record_count'] += 1
        
        if status == '待处理':
            handler_data[handler]['pending_qty'] += quantity
        elif status == '已完成':
            handler_data[handler]['completed_qty'] += quantity
        elif status == '处理中':
            handler_data[handler]['processing_qty'] += quantity
    
    handler_stats = []
    for handler in sorted(handler_data.keys()):
        data = handler_data[handler]
        handler_stats.append({
            'handler': handler,
            'total_qty': data['total_qty'],
            'pending_qty': data['pending_qty'],
            'completed_qty': data['completed_qty'],
            'processing_qty': data['processing_qty'],
            'record_count': data['record_count']
        })
    
    conn.close()
    return jsonify(handler_stats)

@app.route('/api/dashboard/reason-records')
def get_reason_records():
    reason = request.args.get('reason')
    page = int(request.args.get('page', 1))
    per_page = 10
    
    conn = get_db()
    cursor = conn.cursor()
    
    if not reason:
        cursor.execute('''
            SELECT control_reason as reason, 
                   COUNT(*) as count, 
                   COALESCE(SUM(quantity), 0) as quantity
            FROM guankong_records 
            GROUP BY control_reason 
            ORDER BY SUM(quantity) DESC
        ''')
        records = cursor.fetchall()
        conn.close()
        return jsonify([{'reason': r['reason'], 'count': r['count'], 'quantity': r['quantity']} for r in records])
    
    cursor.execute('''
        SELECT COUNT(*) FROM guankong_records 
        WHERE control_reason = ? AND status IN ("待处理", "处理中")
    ''', (reason,))
    total = cursor.fetchone()[0]
    
    offset = (page - 1) * per_page
    cursor.execute('''
        SELECT * FROM guankong_records 
        WHERE control_reason = ? AND status IN ("待处理", "处理中")
        ORDER BY production_date DESC
        LIMIT ? OFFSET ?
    ''', (reason, per_page, offset))
    
    records = cursor.fetchall()
    conn.close()
    
    result = []
    for record in records:
        result.append({
            'id': record['id'],
            'production_date': record['production_date'],
            'product_name': record['product_name'],
            'control_reason': record['control_reason'],
            'quantity': record['quantity'],
            'handled_quantity': record['handled_quantity'] if record['handled_quantity'] is not None else 0,
            'handle_opinion': record['handle_opinion'],
            'handle_time': record['handle_time'],
            'handle_dept': record['handle_dept'],
            'status': record['status'],
            'deadline': record['deadline'],
            'remark': record['remark']
        })
    
    return jsonify({
        'records': result,
        'total': total,
        'page': page,
        'pages': (total + per_page - 1) // per_page
    })

@app.route('/api/overdue_records')
def get_overdue_records():
    conn = get_db()
    cursor = conn.cursor()
    today = datetime.now().strftime('%Y-%m-%d')
    cursor.execute('SELECT * FROM guankong_records WHERE status != "已完成" AND deadline < ? ORDER BY deadline DESC', (today,))
    records = cursor.fetchall()
    conn.close()
    
    result = []
    for record in records:
        result.append({
            'id': record['id'],
            'production_date': record['production_date'],
            'product_name': record['product_name'],
            'control_reason': record['control_reason'],
            'quantity': record['quantity'],
            'deadline': record['deadline']
        })
    return jsonify(result)

@app.route('/api/records', methods=['GET'])
def get_records():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    
    count_query = 'SELECT SUM(quantity), SUM(CASE WHEN handled_quantity IS NOT NULL THEN handled_quantity ELSE 0 END), COUNT(*) FROM guankong_records WHERE 1=1'
    query = 'SELECT * FROM guankong_records WHERE 1=1'
    params = []
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    spec = request.args.get('spec')
    flavor = request.args.get('flavor')
    handle_dept = request.args.get('handle_dept')
    status = request.args.get('status')
    overdue = request.args.get('overdue')
    product_name = request.args.get('product_name')
    deadline = request.args.get('deadline')
    
    if start_date:
        count_query += ' AND production_date >= ?'
        query += ' AND production_date >= ?'
        params.append(start_date)
    if end_date:
        count_query += ' AND production_date <= ?'
        query += ' AND production_date <= ?'
        params.append(end_date)
    if spec and spec != 'all':
        count_query += ' AND product_name LIKE ?'
        query += ' AND product_name LIKE ?'
        params.append(f'%{spec}%')
    if flavor and flavor != 'all':
        count_query += ' AND product_name LIKE ?'
        query += ' AND product_name LIKE ?'
        params.append(f'%{flavor}%')
    if handle_dept and handle_dept != 'all':
        count_query += ' AND handle_dept = ?'
        query += ' AND handle_dept = ?'
        params.append(handle_dept)
    if product_name:
        count_query += ' AND product_name LIKE ?'
        query += ' AND product_name LIKE ?'
        params.append(f'%{product_name}%')
    if deadline:
        count_query += ' AND deadline = ?'
        query += ' AND deadline = ?'
        params.append(deadline)
    
    query += ' ORDER BY (julianday(date(\'now\')) - julianday(deadline)) DESC, deadline ASC'
    
    cursor.execute(query, params)
    records = cursor.fetchall()
    conn.close()
    
    result = []
    total_qty = 0
    handled_qty = 0
    total_count = 0
    
    for record in records:
        quantity = int(record[4])
        handled_quantity = int(record[12]) if record[12] is not None else 0
        db_status = record[8]
        deadline_date = record[9]
        
        display_status = db_status
        if quantity > 0 and handled_quantity > 0 and handled_quantity < quantity:
            display_status = '处理中'
        elif handled_quantity >= quantity > 0:
            display_status = '已完成'
        
        should_include = True
        
        if status and status != 'all':
            if status == '待处理':
                should_include = display_status == '待处理'
            elif status == '处理中':
                should_include = display_status == '处理中'
            elif status == '已完成':
                should_include = display_status == '已完成'
        
        if overdue and overdue != 'all':
            import datetime
            today = datetime.date.today()
            deadline_dt = datetime.datetime.strptime(deadline_date, '%Y-%m-%d').date()
            overdue_days = (today - deadline_dt).days
            
            if overdue == 'not_overdue':
                should_include = overdue_days <= 0
            elif overdue == 'overdue_7':
                should_include = overdue_days > 0 and overdue_days <= 7
            elif overdue == 'overdue_30':
                should_include = overdue_days > 7 and overdue_days <= 30
            elif overdue == 'overdue_more':
                should_include = overdue_days > 30
        
        if should_include:
            total_qty += quantity
            handled_qty += handled_quantity
            total_count += 1
            result.append({
                'id': record[0],
                'production_date': record[1],
                'product_name': record[2],
                'control_reason': record[3],
                'quantity': quantity,
                'handle_opinion': record[5],
                'handle_time': record[6],
                'handle_dept': record[7],
                'status': display_status,
                'deadline': record[9],
                'remark': record[10],
                'created_at': record[11],
                'handled_quantity': handled_quantity,
                'creator_id': record[13],
                'handler_name': record[14] if len(record) > 14 else None
            })
    
    progress = round((handled_qty / total_qty) * 100) if total_qty > 0 else 0
    
    return jsonify({
        'records': result,
        'statistics': {
            'total_qty': total_qty,
            'handled_qty': handled_qty,
            'total_count': total_count,
            'progress': progress
        }
    })

@app.route('/api/records/process', methods=['POST'])
def process_record():
    data = request.json
    record_id = data.get('id')
    quantity = data.get('quantity')
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT quantity, handled_quantity FROM guankong_records WHERE id = ?', (record_id,))
    record = cursor.fetchone()
    
    if record:
        total_qty = int(record[0])
        handled_qty = int(record[1]) if record[1] else 0
        new_handled = handled_qty + quantity
        
        if new_handled >= total_qty:
            cursor.execute('UPDATE guankong_records SET handled_quantity = ?, status = "已完成" WHERE id = ?', (total_qty, record_id))
        else:
            cursor.execute('UPDATE guankong_records SET handled_quantity = ?, status = "处理中" WHERE id = ?', (new_handled, record_id))
        
        conn.commit()
        conn.close()
        return jsonify({'message': f'已处理 {quantity} 件，当前已处理 {new_handled}/{total_qty} 件'})
    
    conn.close()
    return jsonify({'message': '记录不存在'}), 404

@app.route('/api/records/<int:id>', methods=['GET'])
def get_record(id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM guankong_records WHERE id = ?', (id,))
    record = cursor.fetchone()
    conn.close()
    
    if record:
        quantity = int(record[4])
        handled_quantity = int(record[12]) if record[12] else 0
        db_status = record[8]
        
        display_status = db_status
        if quantity > 0 and handled_quantity > 0 and handled_quantity < quantity:
            display_status = '处理中'
        elif handled_quantity >= quantity > 0:
            display_status = '已完成'
        
        return jsonify({
            'id': record[0],
            'production_date': record[1],
            'product_name': record[2],
            'control_reason': record[3],
            'quantity': quantity,
            'handle_opinion': record[5],
            'handle_time': record[6],
            'handle_dept': record[7],
            'status': display_status,
            'deadline': record[9],
            'remark': record[10],
            'created_at': record[11],
            'handled_quantity': handled_quantity,
            'creator_id': record[13],
            'handler': record[14] if len(record) > 14 else None
        })
    
    return jsonify(None)

@app.route('/api/check-duplicate', methods=['GET'])
def check_duplicate():
    production_date = request.args.get('production_date')
    product_name = request.args.get('product_name')
    
    if not production_date or not product_name:
        return jsonify({'exists': False})
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT COUNT(*) FROM guankong_records 
        WHERE production_date = ? AND product_name = ?
    ''', (production_date, product_name))
    
    count = cursor.fetchone()[0]
    conn.close()
    
    return jsonify({'exists': count > 0})

@app.route('/api/records', methods=['POST'])
def add_record():
    data = request.json
    production_date = parse_date(data.get('production_date'))
    
    # 处理期限 = 生产班次 + 7天
    deadline = (datetime.strptime(production_date, '%Y-%m-%d') + timedelta(days=7)).strftime('%Y-%m-%d')
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO guankong_records 
        (production_date, product_name, control_reason, quantity, handle_opinion, handle_time, handle_dept, status, deadline, remark, handled_quantity, creator_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        production_date,
        data.get('product_name'),
        data.get('control_reason'),
        data.get('quantity'),
        data.get('handle_opinion'),
        data.get('handle_time'),
        data.get('handle_dept'),
        data.get('status', '待处理'),
        deadline,
        data.get('remark'),
        data.get('handled_quantity', 0),
        current_user.id
    ))
    conn.commit()
    id = cursor.lastrowid
    conn.close()
    
    return jsonify({'id': id, 'message': 'Record added successfully'}), 201

@app.route('/api/records/<int:id>', methods=['PUT'])
def update_record(id):
    data = request.json
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT quantity, handled_quantity FROM guankong_records WHERE id = ?', (id,))
    current = cursor.fetchone()
    current_qty = current[0] if current else 0
    current_handled = current[1] if current and current[1] else 0
    
    new_handled = int(data.get('handled_quantity', current_handled))
    new_qty = int(data.get('quantity', current_qty))
    
    if 'handled_quantity' in data or 'quantity' in data:
        if new_handled == 0:
            data['status'] = '待处理'
        elif new_handled >= new_qty:
            data['status'] = '已完成'
            if not data.get('handle_time'):
                data['handle_time'] = datetime.now().strftime('%Y-%m-%d')
        else:
            data['status'] = '处理中'
    
    if data.get('status') == '已完成' and not data.get('handle_time'):
        data['handle_time'] = datetime.now().strftime('%Y-%m-%d')
    
    updates = []
    params = []
    
    if 'production_date' in data:
        parsed_date = parse_date(data['production_date'])
        updates.append('production_date = ?')
        params.append(parsed_date)
        # 处理期限 = 生产班次 + 7天
        deadline = (datetime.strptime(parsed_date, '%Y-%m-%d') + timedelta(days=7)).strftime('%Y-%m-%d')
        updates.append('deadline = ?')
        params.append(deadline)
    if 'product_name' in data:
        updates.append('product_name = ?')
        params.append(data['product_name'])
    if 'control_reason' in data:
        updates.append('control_reason = ?')
        params.append(data['control_reason'])
    if 'quantity' in data:
        updates.append('quantity = ?')
        params.append(data['quantity'])
    if 'handled_quantity' in data:
        updates.append('handled_quantity = ?')
        params.append(data['handled_quantity'])
    if 'handle_opinion' in data:
        updates.append('handle_opinion = ?')
        params.append(data['handle_opinion'])
    if 'handle_time' in data:
        updates.append('handle_time = ?')
        params.append(data['handle_time'])
    if 'handle_dept' in data:
        updates.append('handle_dept = ?')
        params.append(data['handle_dept'])
    if 'handler_name' in data:
        updates.append('handler_name = ?')
        params.append(data['handler_name'])
    if 'status' in data:
        updates.append('status = ?')
        params.append(data['status'])
    if 'remark' in data:
        updates.append('remark = ?')
        params.append(data['remark'])
    
    params.append(id)
    
    cursor.execute(f'UPDATE guankong_records SET {", ".join(updates)} WHERE id = ?', params)
    conn.commit()
    conn.close()
    
    return jsonify({'message': 'Record updated successfully'})

@app.route('/api/records/<int:id>', methods=['DELETE'])
def delete_record(id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM guankong_records WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    
    return jsonify({'message': 'Record deleted successfully'})

@app.route('/api/records/delete-all', methods=['DELETE'])
def delete_all_records():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM guankong_records')
    conn.commit()
    deleted_count = cursor.rowcount
    conn.close()
    
    return jsonify({'message': f'已删除 {deleted_count} 条记录'})

@app.route('/api/records/reset-progress', methods=['POST'])
def reset_all_progress():
    """批量重置所有记录的处理进度"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE guankong_records SET handled_quantity = 0, status = "待处理", handle_time = NULL')
    conn.commit()
    updated_count = cursor.rowcount
    conn.close()
    
    return jsonify({'message': f'已重置 {updated_count} 条记录的处理进度'})

@app.route('/api/records/batch-update-dept', methods=['POST'])
def batch_update_dept():
    """批量修改处理部门"""
    data = request.json
    ids = data.get('ids', [])
    dept = data.get('dept', '')
    
    if not ids or not dept:
        return jsonify({'error': '参数错误'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    placeholders = ','.join('?' * len(ids))
    cursor.execute(f'UPDATE guankong_records SET handle_dept = ? WHERE id IN ({placeholders})', [dept] + ids)
    conn.commit()
    updated_count = cursor.rowcount
    conn.close()
    
    return jsonify({'message': f'已修改 {updated_count} 条记录的处理部门'})

@app.route('/api/records/batch-complete', methods=['POST'])
def batch_complete():
    """批量标记为已完成"""
    data = request.json
    ids = data.get('ids', [])
    
    if not ids:
        return jsonify({'error': '参数错误'}), 400
    
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    conn = get_db()
    cursor = conn.cursor()
    placeholders = ','.join('?' * len(ids))
    cursor.execute(f'''
        UPDATE guankong_records 
        SET status = "已完成", handle_time = ?, handled_quantity = quantity 
        WHERE id IN ({placeholders})
    ''', [now] + ids)
    conn.commit()
    updated_count = cursor.rowcount
    conn.close()
    
    return jsonify({'message': f'已标记 {updated_count} 条记录为已完成'})

@app.route('/api/records/batch-delete', methods=['POST'])
def batch_delete():
    """批量删除记录"""
    data = request.json
    ids = data.get('ids', [])
    
    if not ids:
        return jsonify({'error': '参数错误'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    placeholders = ','.join('?' * len(ids))
    cursor.execute(f'DELETE FROM guankong_records WHERE id IN ({placeholders})', ids)
    conn.commit()
    deleted_count = cursor.rowcount
    conn.close()
    
    return jsonify({'message': f'已删除 {deleted_count} 条记录'})

@app.route('/api/records/export')
def export_records():
    """导出选中记录"""
    ids = request.args.get('ids', '')
    if not ids:
        return jsonify({'error': '请选择要导出的记录'}), 400
    
    id_list = [int(x.strip()) for x in ids.split(',') if x.strip()]
    
    conn = get_db()
    cursor = conn.cursor()
    placeholders = ','.join('?' * len(id_list))
    cursor.execute(f'''
        SELECT production_date, product_name, control_reason, quantity, handled_quantity, 
               handle_dept, status, deadline 
        FROM guankong_records 
        WHERE id IN ({placeholders})
    ''', id_list)
    records = cursor.fetchall()
    conn.close()
    
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['生产班次', '产品品项', '管制原因', '数量', '已处理', '处理部门', '状态', '处理期限'])
    
    for record in records:
        writer.writerow([
            record['production_date'],
            record['product_name'],
            record['control_reason'],
            record['quantity'],
            record['handled_quantity'] or 0,
            record['handle_dept'] or '',
            record['status'],
            record['deadline']
        ])
    
    output.seek(0)
    from flask import make_response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename=records_{datetime.now().strftime("%Y%m%d")}.csv'
    
    return response

@app.route('/api/options')
def get_options():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT product_name FROM guankong_records')
    records = cursor.fetchall()
    conn.close()
    
    specs = set()
    flavors = set()
    
    for record in records:
        product_name = record['product_name']
        import re
        match = re.search(r'成品-(PET\d+L?[*\d]*)', product_name)
        if match:
            specs.add(match.group(1))
        match = re.search(r'入(.+)', product_name)
        if match:
            flavor = match.group(1)
            flavor = re.sub(r'(五码合|五码|五码合一版|五码合一|全国五码合一|版|非价签版|彩膜装|透明膜装|纸箱装|无菌|热充|暖饮|天康|津二厂|一厂|果醋版).*', '', flavor).strip()
            flavor = re.sub(r'\(.*\)$', '', flavor).strip()
            flavor = re.sub(r'^[(（].*?[)）]', '', flavor).strip()
            flavors.add(flavor)
    
    return jsonify({
        'specs': sorted(list(specs)),
        'flavors': sorted(list(flavors))
    })

@app.route('/api/export', methods=['GET'])
def export_excel():
    sample = request.args.get('sample')
    
    if sample:
        wb = Workbook()
        ws = wb.active
        ws.title = '管制品数据模板'
        
        headers = ['生产班次', '产品品项', '管制原因', '数量', '处理意见', '处理部门']
        ws.append(headers)
        
        header_fill = PatternFill(start_color='3b82f6', end_color='3b82f6', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        sample_data = [
            ['2026-04-08', '成品-PET330*12入茉莉蜜茶彩膜装（天康）无菌17g', '返工', 240, '返工', '仓储'],
            ['2026-04-07', '成品-PET500*15入绿茶低糖蜂蜜味纸箱装', '条码不符', 800, '更换条码', '制造'],
        ]
        
        for row in sample_data:
            ws.append(row)
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='管制品数据模板_带示例.xlsx'
        )
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = 'SELECT * FROM guankong_records WHERE 1=1'
    params = []
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    spec = request.args.get('spec')
    flavor = request.args.get('flavor')
    handle_dept = request.args.get('handle_dept')
    status = request.args.get('status')
    
    if start_date:
        query += ' AND production_date >= ?'
        params.append(start_date)
    if end_date:
        query += ' AND production_date <= ?'
        params.append(end_date)
    if spec and spec != 'all':
        query += ' AND product_name LIKE ?'
        params.append(f'%{spec}%')
    if flavor and flavor != 'all':
        query += ' AND product_name LIKE ?'
        params.append(f'%{flavor}%')
    if handle_dept and handle_dept != 'all':
        query += ' AND handle_dept = ?'
        params.append(handle_dept)
    if status and status != 'all':
        query += ' AND status = ?'
        params.append(status)
    
    query += ' ORDER BY created_at DESC'
    cursor.execute(query, params)
    records = cursor.fetchall()
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = '管制品记录'
    
    headers = ['ID', '生产班次', '产品品项', '管制原因', '数量', '已处理数量', '处理意见', '处理时间', '处理部门', '状态', '处理期限', '备注', '创建时间']
    ws.append(headers)
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for record in records:
        ws.append([
            record['id'],
            record['production_date'],
            record['product_name'],
            record['control_reason'],
            record['quantity'],
            record['handled_quantity'] if record['handled_quantity'] else 0,
            record['handle_opinion'],
            record['handle_time'] or '未安排',
            record['handle_dept'],
            record['status'],
            record['deadline'],
            record['remark'],
            record['created_at']
        ])
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'管制品记录_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
    )

@app.route('/api/check-template-version')
def check_template_version():
    return jsonify({'match': True})

@app.route('/api/import-preview', methods=['POST'])
def import_preview():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        header_map = {
            '生产班次': 'production_date',
            '产品品项': 'product_name',
            '管制原因': 'control_reason',
            '数量': 'quantity',
            '处理意见': 'handle_opinion',
            '处理时间': 'handle_time',
            '处理部门': 'handle_dept',
            '状态': 'status',
            '已处理数量': 'handled_quantity',
            '处理期限': 'deadline',
            '期限': 'deadline',
            '截止日期': 'deadline',
            '截止期限': 'deadline',
            '规格': 'spec',
            '口味': 'flavor'
        }
        
        rows = []
        pass_count = 0
        warning_count = 0
        error_count = 0
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_data = {'row_number': row_num}
            errors = []
            
            for i, cell in enumerate(row):
                if i < len(headers) and headers[i] in header_map:
                    row_data[header_map[headers[i]]] = cell.value
            
            if 'product_name' in row_data and row_data['product_name']:
                if not row_data.get('production_date'):
                    errors.append({'field': 'production_date', 'message': '生产日期不能为空'})
                
                if not row_data.get('control_reason'):
                    errors.append({'field': 'control_reason', 'message': '管制原因不能为空'})
                
                if not row_data.get('quantity'):
                    errors.append({'field': 'quantity', 'message': '数量不能为空'})
                elif not str(row_data['quantity']).isdigit():
                    errors.append({'field': 'quantity', 'message': '数量必须为整数'})
                
                if not row_data.get('handle_dept'):
                    errors.append({'field': 'handle_dept', 'message': '处理部门不能为空'})
                
                if errors:
                    row_data['status'] = 'error'
                    row_data['error_message'] = '; '.join([e['message'] for e in errors])
                    row_data['suggestion'] = '请检查上述字段并修正后重新上传'
                    error_count += 1
                else:
                    row_data['status'] = 'pass'
                    pass_count += 1
            else:
                if row_data.get('product_name') == '' or row_data.get('product_name') is None:
                    continue
                row_data['status'] = 'warning'
                warning_count += 1
            
            rows.append(row_data)
        
        return jsonify({
            'total': len([r for r in rows if r.get('product_name')]),
            'pass_count': pass_count,
            'warning_count': warning_count,
            'error_count': error_count,
            'rows': rows
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-error-report', methods=['POST'])
def export_error_report():
    try:
        data = request.get_json()
        rows = data.get('rows', [])
        error_rows = [r for r in rows if r.get('status') == 'error']
        
        wb = Workbook()
        ws = wb.active
        ws.title = '导入错误报告'
        
        headers = ['行号', '生产日期', '产品名称', '规格', '口味', '数量', '管制原因', '处理部门', '处理期限', '错误原因', '建议']
        ws.append(headers)
        
        header_fill = PatternFill(start_color='DC3912', end_color='DC3912', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row in error_rows:
            ws.append([
                row.get('row_number'),
                row.get('production_date'),
                row.get('product_name'),
                row.get('spec'),
                row.get('flavor'),
                row.get('quantity'),
                row.get('control_reason'),
                row.get('handle_dept'),
                row.get('deadline'),
                row.get('error_message'),
                row.get('suggestion')
            ])
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 18
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'导入错误报告_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/import-history')
def get_import_history():
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT import_time, operator, filename, total_rows, success_rows, failed_rows, import_type, status 
        FROM import_history 
        ORDER BY import_time DESC
        LIMIT 20
    ''')
    
    records = cursor.fetchall()
    conn.close()
    
    history = []
    for record in records:
        history.append({
            'import_time': record['import_time'],
            'operator': record['operator'],
            'filename': record['filename'],
            'total_rows': record['total_rows'],
            'success_rows': record['success_rows'],
            'failed_rows': record['failed_rows'],
            'import_type': record['import_type'],
            'status': record['status']
        })
    
    return jsonify(history)

def init_import_history_table():
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS import_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_time TEXT,
            operator TEXT,
            filename TEXT,
            total_rows INTEGER,
            success_rows INTEGER,
            failed_rows INTEGER,
            import_type TEXT,
            status TEXT
        )
    ''')
    
    conn.commit()
    conn.close()

@app.route('/api/import', methods=['POST'])
def import_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        header_map = {
            '生产班次': 'production_date',
            '产品品项': 'product_name',
            '管制原因': 'control_reason',
            '数量': 'quantity',
            '处理意见': 'handle_opinion',
            '处理时间': 'handle_time',
            '处理部门': 'handle_dept',
            '状态': 'status',
            '已处理数量': 'handled_quantity',
            '处理期限': 'deadline',
            '期限': 'deadline',
            '截止日期': 'deadline',
            '截止期限': 'deadline'
        }
        
        conn = get_db()
        cursor = conn.cursor()
        deadline_days = int(get_config('default_deadline_days') or 2)
        
        total_rows = 0
        for row in ws.iter_rows(min_row=2):
            row_data = {}
            for i, cell in enumerate(row):
                if i < len(headers) and headers[i] in header_map:
                    row_data[header_map[headers[i]]] = cell.value
            
            if 'product_name' in row_data and row_data['product_name']:
                total_rows += 1
        
        imported_count = 0
        for row in ws.iter_rows(min_row=2):
            row_data = {}
            for i, cell in enumerate(row):
                if i < len(headers) and headers[i] in header_map:
                    row_data[header_map[headers[i]]] = cell.value
            
            if 'product_name' in row_data and row_data['product_name']:
                production_date = parse_date(row_data.get('production_date'))
                
                # 优先使用上传数据中的处理期限，否则默认7天
                deadline_value = row_data.get('deadline')
                if deadline_value is not None and str(deadline_value).strip():
                    deadline = parse_date(deadline_value)
                else:
                    deadline = (datetime.strptime(production_date, '%Y-%m-%d') + timedelta(days=7)).strftime('%Y-%m-%d')
                
                handle_time = parse_date(row_data.get('handle_time'))
                
                cursor.execute('''
                    INSERT INTO guankong_records 
                    (production_date, product_name, control_reason, quantity, handle_opinion, handle_time, handle_dept, status, deadline, handled_quantity, creator_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    production_date,
                    row_data.get('product_name'),
                    row_data.get('control_reason', ''),
                    int(row_data.get('quantity', 0)),
                    row_data.get('handle_opinion', ''),
                    handle_time if handle_time != datetime.now().strftime('%Y-%m-%d') else None,
                    row_data.get('handle_dept', ''),
                    row_data.get('status', '待处理'),
                    deadline,
                    int(row_data.get('handled_quantity', 0)),
                    current_user.id
                ))
                imported_count += 1
        
        conn.commit()
        
        cursor.execute('''
            INSERT INTO import_history 
            (import_time, operator, filename, total_rows, success_rows, failed_rows, import_type, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            current_user.username if current_user else '系统用户',
            file.filename,
            total_rows,
            imported_count,
            total_rows - imported_count,
            '新增',
            'success' if (total_rows - imported_count) == 0 else 'partial'
        ))
        conn.commit()
        conn.close()
        
        return jsonify({'success_count': imported_count}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/daily_report')
def get_daily_report():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    
    today = datetime.now().strftime('%Y-%m-%d')
    
    cursor.execute('SELECT COUNT(*), SUM(quantity) FROM guankong_records WHERE status = "待处理"')
    pending_data = cursor.fetchone()
    if pending_data:
        pending_count = pending_data[0]
        pending_qty = pending_data[1] or 0
    else:
        pending_count = 0
        pending_qty = 0
    
    cursor.execute('SELECT COUNT(*), SUM(quantity) FROM guankong_records WHERE status = "已完成" AND DATE(handle_time) = ?', (today,))
    completed_data = cursor.fetchone()
    if completed_data:
        completed_count = completed_data[0]
        completed_qty = completed_data[1] or 0
    else:
        completed_count = 0
        completed_qty = 0
    
    cursor.execute('SELECT COUNT(*), SUM(quantity) FROM guankong_records WHERE status = "处理中"')
    processing_data = cursor.fetchone()
    if processing_data:
        processing_count = processing_data[0]
        processing_qty = processing_data[1] or 0
    else:
        processing_count = 0
        processing_qty = 0
    
    cursor.execute('''
        SELECT COUNT(*) FROM guankong_records 
        WHERE status = "待处理" AND deadline IS NOT NULL AND deadline < date('now')
    ''')
    overdue_count = cursor.fetchone()[0] or 0
    
    cursor.execute('''
        SELECT product_name, COUNT(*), SUM(quantity), AVG(julianday('now') - julianday(deadline)) 
        FROM guankong_records 
        WHERE status = "待处理" AND deadline IS NOT NULL
        GROUP BY product_name
        ORDER BY SUM(quantity) DESC
    ''')
    pending_products = []
    for row in cursor.fetchall():
        pending_products.append({
            'name': row[0],
            'count': row[1],
            'quantity': row[2],
            'avg_overdue_days': max(0, round(row[3])) if row[3] else 0
        })
    
    cursor.execute('SELECT product_name, COUNT(*), SUM(quantity) FROM guankong_records WHERE status = "待处理" GROUP BY product_name')
    pending_by_product = cursor.fetchall()
    
    cursor.execute('SELECT product_name, COUNT(*), SUM(quantity) FROM guankong_records WHERE status = "已完成" AND DATE(handle_time) = ? GROUP BY product_name', (today,))
    completed_by_product = cursor.fetchall()
    
    cursor.execute('SELECT product_name, quantity, deadline FROM guankong_records WHERE status = "待处理" ORDER BY created_at DESC')
    pending_records = cursor.fetchall()
    
    cursor.execute('SELECT product_name, handled_quantity, handle_time FROM guankong_records WHERE status = "已完成" AND DATE(handle_time) = ? ORDER BY handle_time DESC', (today,))
    today_completed = cursor.fetchall()
    
    cursor.execute('''
        SELECT DATE(handle_time) as date, COUNT(*) as count 
        FROM guankong_records 
        WHERE status = "已完成" AND handle_time IS NOT NULL
        GROUP BY DATE(handle_time)
        ORDER BY DATE(handle_time) DESC
        LIMIT 10
    ''')
    completed_dates = [row[0] for row in cursor.fetchall()]
    
    no_progress_days = 0
    check_date = datetime.now()
    while no_progress_days < 30:
        check_date_str = check_date.strftime('%Y-%m-%d')
        if check_date_str in completed_dates:
            break
        no_progress_days += 1
        check_date -= timedelta(days=1)
    
    conn.close()
    
    report = f"""【管制品追踪日报】
日期：{today}

一、待处理情况
总待处理记录数：{pending_count} 条
总待处理数量：{pending_qty} 件
超期记录数：{overdue_count} 条

按产品分类：
"""
    for row in pending_by_product:
        report += f"  - {row[0]}：{row[1]}条记录，{row[2]}件\n"
    
    report += f"""

二、今日已处理情况
今日已完成记录数：{completed_count} 条
今日已完成数量：{completed_qty} 件

按产品分类：
"""
    for row in completed_by_product:
        report += f"  - {row[0]}：{row[1]}条记录，{row[2]}件\n"
    
    report += f"""

三、处理中情况
处理中记录数：{processing_count} 条
处理中数量：{processing_qty} 件

===============================
统计时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""
    
    pending_records_list = [
        {'product_name': row[0], 'quantity': row[1], 'deadline': row[2]} 
        for row in pending_records
    ]
    
    today_completed_list = [
        {'product_name': row[0], 'handled_quantity': row[1], 'handle_time': row[2], 'handler': None} 
        for row in today_completed
    ]
    
    return jsonify({
        'report': report,
        'pending_count': pending_count,
        'pending_qty': pending_qty,
        'completed_count': completed_count,
        'completed_qty': completed_qty,
        'processing_count': processing_count,
        'processing_qty': processing_qty,
        'overdue_count': overdue_count,
        'pending_products': pending_products,
        'pending_by_product': [{'name': row[0], 'count': row[1], 'quantity': row[2]} for row in pending_by_product],
        'completed_by_product': [{'name': row[0], 'count': row[1], 'quantity': row[2]} for row in completed_by_product],
        'pending_records': pending_records_list,
        'today_completed': today_completed_list,
        'no_progress_days': no_progress_days
    })

@app.route('/api/report_data', methods=['GET'])
def get_report_data():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT COALESCE(control_reason, '未分类') as reason, 
               SUM(quantity) as total_qty, 
               COUNT(*) as total_count
        FROM guankong_records 
        WHERE strftime("%Y-%m", production_date) = ?
        GROUP BY COALESCE(control_reason, '未分类')
        ORDER BY SUM(quantity) DESC
    ''', (month,))
    reason_stats = []
    for row in cursor.fetchall():
        reason_stats.append({'reason': row[0], 'quantity': row[1] or 0, 'count': row[2]})
    
    if not reason_stats:
        cursor.execute('''
            SELECT COALESCE(control_reason, '未分类') as reason, 
                   SUM(quantity) as total_qty, 
                   COUNT(*) as total_count
            FROM guankong_records 
            GROUP BY COALESCE(control_reason, '未分类')
            ORDER BY SUM(quantity) DESC
        ''')
        for row in cursor.fetchall():
            reason_stats.append({'reason': row[0], 'quantity': row[1] or 0, 'count': row[2]})
    
    # print(f"DEBUG: Month={month}, Reason stats count={len(reason_stats)}, first item: {reason_stats[0] if reason_stats else 'None'}")
    
    cursor.execute('''
        SELECT strftime("%Y-%m", production_date) as month, COUNT(*) as count 
        FROM guankong_records 
        WHERE production_date IS NOT NULL
        GROUP BY strftime("%Y-%m", production_date)
        ORDER BY month DESC
        LIMIT 6
    ''')
    monthly_data = []
    for row in cursor.fetchall():
        monthly_data.append({'month': row[0], 'count': row[1]})
    
    current_month_count = 0
    last_month_count = 0
    current_month = datetime.now().strftime('%Y-%m')
    last_month = (datetime.now() - timedelta(days=30)).strftime('%Y-%m')
    
    for item in monthly_data:
        if item['month'] == current_month:
            current_month_count = item['count']
        elif item['month'] == last_month:
            last_month_count = item['count']
    
    monthly_trend = 0
    if last_month_count > 0:
        monthly_trend = round(((current_month_count - last_month_count) / last_month_count) * 100)
    
    cursor.execute('''
        SELECT COALESCE(handler_name, handle_dept) as dept, AVG(julianday(handle_time) - julianday(production_date)) * 24, COUNT(*) 
        FROM guankong_records 
        WHERE status = "已完成" AND handle_time IS NOT NULL AND production_date IS NOT NULL
        AND strftime("%Y-%m", production_date) = ?
        AND COALESCE(handler_name, handle_dept) IS NOT NULL
        AND COALESCE(handler_name, handle_dept) != ""
        AND COALESCE(handler_name, handle_dept) != "未分配"
        GROUP BY COALESCE(handler_name, handle_dept)
    ''', (month,))
    avg_duration = []
    for row in cursor.fetchall():
        avg_duration.append({'dept': row[0], 'hours': round(row[1], 1), 'count': row[2]})
    
    if not avg_duration:
        cursor.execute('''
            SELECT COALESCE(handler_name, handle_dept) as dept, AVG(julianday(handle_time) - julianday(production_date)) * 24, COUNT(*) 
            FROM guankong_records 
            WHERE status = "已完成" AND handle_time IS NOT NULL AND production_date IS NOT NULL
            AND COALESCE(handler_name, handle_dept) IS NOT NULL
            AND COALESCE(handler_name, handle_dept) != ""
            AND COALESCE(handler_name, handle_dept) != "未分配"
            GROUP BY COALESCE(handler_name, handle_dept)
        ''')
        for row in cursor.fetchall():
            avg_duration.append({'dept': row[0], 'hours': round(row[1], 1), 'count': row[2]})
    
    cursor.execute('SELECT COUNT(*) FROM guankong_records')
    result = cursor.fetchone()
    total_count = result[0] if result else 0
    
    cursor.execute('SELECT SUM(quantity) FROM guankong_records WHERE status = "待处理"')
    result = cursor.fetchone()
    pending_qty = result[0] if result and result[0] else 0
    
    cursor.execute('SELECT COUNT(*) FROM guankong_records WHERE status = "待处理"')
    result = cursor.fetchone()
    pending_count = result[0] if result else 0
    
    cursor.execute('SELECT COALESCE(handler_name, handle_dept) as dept, COALESCE(SUM(quantity), 0) FROM guankong_records WHERE status = "待处理" AND COALESCE(handler_name, handle_dept) IS NOT NULL AND COALESCE(handler_name, handle_dept) != "" AND COALESCE(handler_name, handle_dept) != "未分配" GROUP BY COALESCE(handler_name, handle_dept)')
    pending_by_dept = {}
    for row in cursor.fetchall():
        pending_by_dept[row[0]] = row[1]
    
    cursor.execute('''
        SELECT COUNT(*) FROM guankong_records 
        WHERE status = "待处理" AND deadline IS NOT NULL AND deadline < date('now')
    ''')
    result = cursor.fetchone()
    overdue_count = result[0] if result else 0
    
    cursor.execute('''
        SELECT SUM(quantity) FROM guankong_records 
        WHERE status = "待处理" AND deadline IS NOT NULL AND deadline < date('now')
    ''')
    result = cursor.fetchone()
    overdue_qty = result[0] if result and result[0] else 0
    
    cursor.execute('''
        SELECT AVG(julianday(handle_time) - julianday(production_date)) * 24 
        FROM guankong_records 
        WHERE status = "已完成" AND handle_time IS NOT NULL AND production_date IS NOT NULL
    ''')
    result = cursor.fetchone()
    avg_duration_hours = result[0] if result else None
    if avg_duration_hours:
        avg_duration_hours = round(avg_duration_hours, 1)
    
    target_hours = 48
    
    dept_percentages = {}
    for dept, qty in pending_by_dept.items():
        dept_percentages[dept] = round((qty / pending_qty) * 100) if pending_qty > 0 else 0
    overdue_percent = round((overdue_count / pending_count) * 100) if pending_count > 0 else 0
    
    conn.close()
    
    return jsonify({
        'reason_stats': reason_stats, 
        'avg_duration': avg_duration,
        'total_count': total_count,
        'pending_qty': pending_qty,
        'pending_count': pending_count,
        'pending_by_dept': pending_by_dept,
        'dept_percentages': dept_percentages,
        'overdue_count': overdue_count,
        'overdue_qty': overdue_qty,
        'overdue_percent': overdue_percent,
        'avg_duration_hours': avg_duration_hours,
        'target_hours': target_hours,
        'monthly_data': monthly_data,
        'monthly_trend': monthly_trend
    })

@app.route('/api/reason-details', methods=['GET'])
def get_reason_details():
    reason = request.args.get('reason')
    
    if not reason:
        return jsonify({'error': '缺少参数'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT product_name, SUM(quantity) as total_qty, COUNT(*) as count
        FROM guankong_records 
        WHERE control_reason = ?
        GROUP BY product_name
        ORDER BY total_qty DESC
        LIMIT 3
    ''', (reason,))
    
    products = []
    total_records = 0
    total_quantity = 0
    
    for row in cursor.fetchall():
        products.append({
            'name': row[0],
            'quantity': row[1],
            'count': row[2]
        })
        total_records += row[2]
        total_quantity += row[1]
    
    conn.close()
    
    return jsonify({
        'reason': reason,
        'products': products,
        'total_records': total_records,
        'total_quantity': total_quantity
    })

@app.route('/api/users')
def get_users():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT id, username, role, department FROM users')
    users = []
    for row in cursor.fetchall():
        users.append({'id': row['id'], 'username': row['username'], 'role': row['role'], 'department': row['department']})
    conn.close()
    
    return jsonify(users)

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    work_id = data.get('work_id')
    password = data.get('password')
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users WHERE username = ? AND password = ?', (work_id, password))
    user = cursor.fetchone()
    conn.close()
    
    if user:
        return jsonify({
            'token': 'mock_jwt_token_for_' + work_id,
            'user': {
                'id': user['id'],
                'username': user['username'],
                'role': user['role'],
                'department': user['department']
            }
        })
    else:
        return jsonify({'error': '工号或密码错误'}), 401

@app.route('/api/users', methods=['POST'])
def add_user():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    role = data.get('role', '操作员')
    department = data.get('department', '仓储')
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute('INSERT INTO users (username, password, role, department) VALUES (?, ?, ?, ?)',
                      (username, password, role, department))
        conn.commit()
        conn.close()
        return jsonify({'message': 'User added successfully'}), 201
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': '用户名已存在'}), 400

@app.route('/api/users/<int:id>', methods=['PUT'])
def update_user(id):
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    updates = []
    params = []
    
    if 'role' in data:
        updates.append('role = ?')
        params.append(data['role'])
    if 'department' in data:
        updates.append('department = ?')
        params.append(data['department'])
    if 'password' in data and data['password']:
        updates.append('password = ?')
        params.append(data['password'])
    
    params.append(id)
    
    cursor.execute(f'UPDATE users SET {", ".join(updates)} WHERE id = ?', params)
    conn.commit()
    conn.close()
    
    return jsonify({'message': 'User updated successfully'})

@app.route('/api/users/<int:id>', methods=['DELETE'])
def delete_user(id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM users WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    
    return jsonify({'message': 'User deleted successfully'})

@app.route('/api/system_config')
def get_system_config():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT config_key, config_value, description FROM system_config')
    config = {}
    for row in cursor.fetchall():
        config[row['config_key']] = {'value': row['config_value'], 'description': row['description']}
    conn.close()
    
    return jsonify(config)

@app.route('/api/system_config', methods=['POST'])
def update_system_config():
    data = request.json
    for key, value in data.items():
        set_config(key, value)
    
    return jsonify({'message': '配置更新成功'})

# 初始化数据库（云端部署时自动执行）
init_db()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port, use_reloader=False)